import NiaDic
import openpyxl
from tkinter import filedialog

dic = NiaDic.NiaDic()

def splitword(word, onlynoun=True):
    res = []
    if len(word) < 2:
        return [[[word,'unk']]]
    searchres = dic.findword(word)
    if searchres[1] != 'unk' and (not onlynoun or 'nc' in searchres[1]):
        return [[searchres]]

    isanswer = False
    for i in range(len(word)-1):
        suffix = word[-i-1:]
        searchres = dic.findword(suffix)
        if searchres[1] != 'unk' and (not onlynoun or 'nc' in searchres[1]):
            prefix = word[:-i-1]
            if prefix != '':
                prevres = splitword(prefix)
                for resu in prevres:
                    resu.append(searchres)
                    res.append(resu)
                    isanswer = True
    if not isanswer:
        return [[[word,'unk']]]
    return res

def getsplit(word):
    results = splitword(word)
    results2 = []
    for result in results:
        unknum = 0
        solonum = 0
        for word in result:
            if word[1] == 'unk':
                unknum += 1
            if len(word[0]) == 1:
                solonum += 1
        results2.append([result, len(result), unknum, solonum])
    results2.sort(key=lambda word: word[3], reverse=True)
    results2.sort(key=lambda word: word[2], reverse=True)
    results2.sort(key=lambda word: word[1], reverse=True)
    print(results2[-1])
    return results2[-1]

def readxl(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    wordlist = []
    colA = ws['A']
    for word in colA:
        wordlist.append(str(word.value))
    print(f'{len(wordlist)}개의 단어 입력받음')
    wb.close()
    return wordlist

def writexl(wordlist, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    i = 0
    for word in wordlist:
        i+=1
        splited = getsplit(word)
        ws.cell(row = i, column= 1, value=word)
        for j in range(len(splited[0])):
            ws.cell(row=i,column=j+3, value=splited[0][j][0])
            j+=1
    wb.save(filename)
    wb.close()

filename = filedialog.askopenfile(filetypes=(("XLSX files","*.xlsx"), ("All files","*.*"))).name
dest_filename = str(filename).split('/')[-1].split('.')[0]+"_result.xlsx"
writexl(readxl(filename), dest_filename)



